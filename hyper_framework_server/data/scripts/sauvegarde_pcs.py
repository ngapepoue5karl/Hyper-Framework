# sauvegarde_onedrive_hyper.py
__hyper_inputs__ = [
    {"key": "onedrive_file", "label": "OneDriveUsageAccountDetail (.csv)"},
    {"key": "users_file",  "label": "Users (.csv)"},
    {"key": "ad_file",     "label": "Utilisateur AD (.txt)"}
]

import pandas as pd
import numpy as np
import os
import re
from datetime import date
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ================================
# Utilitaires de lecture/clean
# ================================
def read_csv_safely(path: str) -> pd.DataFrame:
    for enc in ["utf-8-sig", "latin-1"]:
        try:
            df = pd.read_csv(path, encoding=enc, low_memory=False)
            df.columns = [c.strip().replace("\ufeff", "") for c in df.columns]
            return df
        except Exception:
            continue
    df = pd.read_csv(path, low_memory=False)
    df.columns = [c.strip().replace("\ufeff", "") for c in df.columns]
    return df

def read_ad_utf16_mixed(path: str) -> pd.DataFrame:
    with open(path, "r", encoding="utf-16", errors="replace") as f:
        lines = [ln.strip("\n\r") for ln in f.readlines()]
    header_line = None
    for ln in lines:
        if ln and not ln.startswith("-") and ln.count(",") >= 20:
            header_line = ln
            break
    if not header_line:
        raise RuntimeError("Header introuvable dans le fichier AD")
    header = [h.strip() for h in header_line.split(",")]
    start = lines.index(header_line) + 1
    rows = []
    for ln in lines[start:]:
        if not ln or ln.startswith("-"):
            continue
        parts = [p.strip() for p in ln.split(";")]
        if len(parts) < len(header):
            parts += [""] * (len(header) - len(parts))
        elif len(parts) > len(header):
            parts = parts[:len(header)]
        rows.append(parts)
    df = pd.DataFrame(rows, columns=header)
    df.columns = [c.strip() for c in df.columns]
    return df

def find_col(df: pd.DataFrame, candidates) -> str | None:
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    for c in df.columns:
        if any(cand.lower() in c.lower() for cand in candidates):
            return c
    return None

_ILLEGAL_CTRL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
def sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Optimisé pour améliorer les performances"""
    out = df.copy()
    # Traiter uniquement les colonnes de type objet (string)
    object_cols = out.select_dtypes(include=["object"]).columns
    for c in object_cols:
        # Vectoriser le nettoyage des caractères illégaux
        out[c] = out[c].astype(str).str.replace(_ILLEGAL_CTRL_CHARS, "", regex=True)
        out[c] = out[c].replace({"nan": None, "None": None})
    return out

def dataframe_for_ui(df: pd.DataFrame) -> pd.DataFrame:
    """Optimisé pour éviter les copies inutiles et accélérer la conversion"""
    df2 = df.replace({pd.NaT: None})
    df2 = df2.where(pd.notnull(df2), None)

    # Conversion optimisée des types 
    for col in df2.columns:
        if df2[col].dtype.kind in ('i', 'u'):  # entiers
            df2[col] = df2[col].astype(object).where(df2[col].notna(), None)
        elif df2[col].dtype.kind == 'f':  # flottants
            df2[col] = df2[col].astype(object).where(df2[col].notna(), None)
        elif df2[col].dtype.kind == 'b':  # booléens
            df2[col] = df2[col].astype(object).where(df2[col].notna(), None)

    return df2

# ================================
# Formules Excel dans "NOK OneDrive"
# ================================
def inject_formulas_nok(writer):
    NOK = "NOK OneDrive"
    LIC = "Liste des licences OneDrive"
    EXT = "Extraction OneDrive"
    ws = writer.sheets[NOK]

    headers = [
        "Service/générique?",
        "Compte Expiré?",
        "Comptes ayant une licence OneDrive",
        "Comptes Sauvegardés",
        "Date d'obtention de la licence",
        "Date dernière synchronisation",
        ">= 30 Jrs"
    ]
    base = ws.max_column
    for j, h in enumerate(headers, 1):
        ws.cell(1, base + j, h)

    COL_A, COL_C, COL_E, COL_G, COL_M, COL_S, COL_W, COL_Y = "A","C","E","G","M","S","W","Y"
    col_srv, col_exp, col_haslic, col_saved, col_lic, col_sync, col_30 = [
        get_column_letter(base+i) for i in range(1,8)
    ]

    def f_service(r): return (
        f'=OR(ISNUMBER(SEARCH("DefaultAccount",{COL_A}{r})),ISNUMBER(SEARCH("generique",{COL_C}{r})),'
        f'ISNUMBER(SEARCH("compte sce",{COL_C}{r})),ISNUMBER(SEARCH("BAL partagee",{COL_C}{r})),'
        f'ISNUMBER(SEARCH("administrator",{COL_W}{r})),ISNUMBER(SEARCH("fournisseur",{COL_W}{r})),'
        f'ISNUMBER(SEARCH("BAL partagee",{COL_W}{r})),ISNUMBER(SEARCH("generique",{COL_G}{r})),'
        f'ISNUMBER(SEARCH("compte sce",{COL_G}{r})),ISNUMBER(SEARCH("BAL partagee",{COL_G}{r})),'
        f'ISNUMBER(SEARCH("sabc",{COL_S}{r})))'
    )
    def f_expire(r):
        acc=f"{COL_Y}{r}"
        dateconv=f"DATE(RIGHT(LEFT({acc},10),4),LEFT({acc},2),MID(LEFT({acc},10),4,2))"
        return f'=IF(AND({acc}<>"",IFERROR({dateconv},"")<Stats!$B$2),"OUI","NON")'
    def f_haslic(r):
        v=f"VLOOKUP({COL_S}{r},'{LIC}'!C:S,17,FALSE)"
        return f'=IF(ISNUMBER(SEARCH("365",{v})),"OUI","NON")'
    def f_saved(r):
        v=f"VLOOKUP({COL_S}{r},'{EXT}'!K:K,1,FALSE)"
        return f'=IF(ISERROR({v}),"NON","OUI")'
    def f_date_lic(r):
        v=f"VLOOKUP({COL_S}{r},'{LIC}'!C:G,5,FALSE)"
        return f'=IFERROR({v},"")'
    def f_last_sync(r):
        v=f"VLOOKUP({COL_S}{r},'{LIC}'!C:Q,15,FALSE)"
        return f'=IFERROR(IF({v}<>"",LEFT({v},LEN({v})-10),""),"")'
    for r in range(2, ws.max_row+1):
        ws[f"{col_srv}{r}"].value   = f_service(r)
        ws[f"{col_exp}{r}"].value   = f_expire(r)
        ws[f"{col_haslic}{r}"].value= f_haslic(r)
        ws[f"{col_saved}{r}"].value = f_saved(r)
        ws[f"{col_lic}{r}"].value   = f_date_lic(r)
        ws[f"{col_sync}{r}"].value  = f_last_sync(r)
    for r in range(2, ws.max_row+1):
        ws[f"{col_30}{r}"].value=(
            f'=IF(AND(TRIM(E{r})="True",TRIM(M{r})="False",'
            f'TRIM({col_srv}{r})="FAUX",TRIM({col_exp}{r})="NON",'
            f'TRIM({col_haslic}{r})="OUI",TRIM({col_saved}{r})="OUI"),'
            f'"OK",IFERROR(IF((Stats!$B$2-{col_sync}{r})>30,"NOK","OK"),""))'
        )

    # Mettre la feuille en Tableau
    last_col = get_column_letter(ws.max_column)
    ref = f"A1:{last_col}{ws.max_row}"
    tbl = Table(displayName="Table_NOK", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tbl.tableStyleInfo = style
    if "Table_NOK" in ws.tables:
        del ws.tables["Table_NOK"]
    ws.add_table(tbl)

# ================================
# Duplication F -> M (Extraction OneDrive)
# ================================
def duplicate_last_activity_column(writer):
    ws = writer.sheets["Extraction OneDrive"]
    if ws.max_column < 6:
        return
    ws.cell(row=1, column=13, value=ws.cell(row=1, column=6).value or "Last Activity Date")
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=13, value=ws.cell(row=r, column=6).value)

# ================================
# Heuristiques pour l'affichage (calcul filtres côté app)
# ================================
def parse_date_safe(v):
    if pd.isna(v) or str(v).strip() == "" or str(v).strip() == "0":
        return pd.NaT
    for fmt in (None, "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            if fmt is None:
                return pd.to_datetime(v, errors="raise")
            else:
                return pd.to_datetime(str(v), format=fmt, errors="raise")
        except Exception:
            continue
    return pd.to_datetime(v, errors="coerce")

def compute_app_flags(ad_df: pd.DataFrame, users_df: pd.DataFrame, onedrive_df: pd.DataFrame) -> pd.DataFrame:
    df = ad_df.copy()

    # Colonnes de base
    col_enabled    = find_col(df, ["Enabled"]) or "Enabled"
    col_pwdexpired = find_col(df, ["PasswordExpired"]) or "PasswordExpired"
    df[col_enabled]    = df[col_enabled].astype(str).str.strip()
    df[col_pwdexpired] = df[col_pwdexpired].astype(str).str.strip()

    col_upn = find_col(df, ["UserPrincipalName","UPN","Email"]) or "UserPrincipalName"

    # Service/générique?
    parts = []
    for c in ["CN","Description","Title","Department","Office","DisplayName"]:
        cc = find_col(df, [c])
        if cc: parts.append(df[cc].astype(str))
    hay = pd.Series([""] * len(df)) if not parts else parts[0]
    for p in parts[1:]:
        hay = hay.str.cat(p.astype(str), sep=" ")
    kw = ["defaultaccount","generique","compte sce","bal partagee","administrator","fournisseur","sabc"]
    is_service = hay.str.lower().apply(lambda s: any(k in s for k in kw))
    df["Service/générique?"] = np.where(is_service, "VRAI", "FAUX")

    # Compte Expiré?
    col_acc = find_col(df, ["AccountExpirationDate"]) or "AccountExpirationDate"
    dexp = df[col_acc].apply(lambda x: parse_date_safe(str(x)[:10] if x is not None else x))
    today = pd.Timestamp(date.today())
    df["Compte Expiré?"] = np.where((~dexp.isna()) & (dexp < today), "OUI", "NON")

    # Licences OneDrive (presence de "365" dans Users[*])
    u_upn = find_col(users_df, ["UserPrincipalName","UPN","User","Email","OwnerPrincipalName","UserId"]) or users_df.columns[0]
    users_df["_row_has_365_"] = users_df.astype(str).apply(lambda s: s.str.contains("365", case=False, na=False)).any(axis=1)
    lic_map = users_df.groupby(u_upn)["_row_has_365_"].max()
    df["Comptes ayant une licence OneDrive"] = df[col_upn].map(lambda u: "OUI" if lic_map.get(str(u), False) else "NON")

    # Comptes Sauvegardés (présence UPN dans Extraction OneDrive col K par défaut)
    ext_upn_col = find_col(onedrive_df, ["UserPrincipalName","OwnerPrincipalName","UserId","Email"]) or onedrive_df.columns[min(10, len(onedrive_df.columns)-1)]
    saved_set = set(onedrive_df[ext_upn_col].astype(str).str.lower())
    df["Comptes Sauvegardés"] = df[col_upn].astype(str).str.lower().isin(saved_set)
    df["Comptes Sauvegardés"] = df["Comptes Sauvegardés"].map(lambda v: "OUI" if v else "NON")

    # Date d'obtention de la licence (colonne G de Users = 5e colonne depuis C)
    if users_df.shape[1] >= 7:  
        date_lic_col = users_df.columns[6]  
        date_lic_map = users_df.set_index(u_upn)[date_lic_col] if u_upn in users_df.columns else pd.Series(dtype=object)
        df["Date d'obtention de la licence"] = df[col_upn].map(lambda u: date_lic_map.get(u, ""))
    else:
        df["Date d'obtention de la licence"] = ""

    # Date dernière synchronisation (si utile)
    if users_df.shape[1] >= 17:
        last_sync_col = users_df.columns[16]  # approx C:Q -> 17e col absolue
        raw_last_sync = users_df.set_index(u_upn)[last_sync_col] if u_upn in users_df.columns else pd.Series(dtype=object)
        def trim10(x):
            s = str(x) if not pd.isna(x) else ""
            return s[:-10] if len(s) >= 10 else ""
        df["Date dernière synchronisation"] = df[col_upn].map(lambda u: trim10(raw_last_sync.get(u, "")))
    else:
        df["Date dernière synchronisation"] = ""

    # >= 30 Jrs (cohérent avec Excel)
    def status_row(row):
        en = str(row.get(col_enabled)).strip().lower() == "true"
        pw = str(row.get(col_pwdexpired)).strip().lower() == "false"
        srv = str(row.get("Service/générique?")).strip().upper() == "FAUX"
        exp = str(row.get("Compte Expiré?")).strip().upper() == "NON"
        lic = str(row.get("Comptes ayant une licence OneDrive")).strip().upper() == "OUI"
        sav = str(row.get("Comptes Sauvegardés")).strip().upper() == "OUI"
        if en and pw and srv and exp and lic and sav:
            return "OK"
        d = parse_date_safe(row.get("Date dernière synchronisation"))
        if not pd.isna(d):
            return "NOK" if (today - d).days > 30 else "OK"
        return ""
    df[">= 30 Jrs"] = df.apply(status_row, axis=1)

    return df

# ================================
# Traitement principal
# ================================
def traiter(onedrive_path, users_path, ad_path, output_dir):
    onedrive_df = read_csv_safely(onedrive_path)   # Extraction OneDrive
    users_df    = read_csv_safely(users_path)      # Liste des licences OneDrive
    ad_df       = read_ad_utf16_mixed(ad_path)     # Utilisateur AD
    nok         = ad_df.copy()                     # base pour NOK OneDrive

    out_xlsx = os.path.join(output_dir, "Rapport de conformité sauvegarde_Pcs.xlsx")

    with pd.ExcelWriter(out_xlsx, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        sanitize_df_for_excel(users_df).to_excel(writer, index=False, sheet_name="Liste des licences OneDrive")
        sanitize_df_for_excel(onedrive_df).to_excel(writer, index=False, sheet_name="Extraction OneDrive")
        sanitize_df_for_excel(ad_df).to_excel(writer, index=False, sheet_name="Utilisateur AD")
        sanitize_df_for_excel(nok).to_excel(writer, index=False, sheet_name="NOK OneDrive")

        # Stats (B2 = date du jour en date Excel)
        ws_stats = writer.book.create_sheet("Stats")
        ws_stats["A1"] = "Indicateur"
        ws_stats["B1"] = "Valeur"
        ws_stats["A2"] = "Date du jour (Stats!B2)"
        ws_stats["B2"].value = date.today()
        ws_stats["B2"].number_format = "yyyy-mm-dd"

        duplicate_last_activity_column(writer)
        inject_formulas_nok(writer)

    # ============ AFFICHAGE APPLICATION ============
    app_df = compute_app_flags(ad_df, users_df, onedrive_df)

    # Colonnes
    cn_col      = find_col(app_df, ["CN", "Name", "DisplayName"]) or "CN"
    upn_col     = find_col(app_df, ["UserPrincipalName", "UPN", "Email"]) or "UserPrincipalName"
    desc_col    = find_col(app_df, ["Description"]) or "Description"
    dept_col    = find_col(app_df, ["Department", "Service"]) or "Department"
    title_col   = find_col(app_df, ["Title", "JobTitle"]) or "Title"
    office_col  = find_col(app_df, ["Office", "Location"]) or "Office"
    enabled_col = find_col(app_df, ["Enabled"]) or "Enabled"
    pwd_col     = find_col(app_df, ["PasswordExpired"]) or "PasswordExpired"

    # Filtres demandés pour la vue
    mask = (
        (app_df[enabled_col].astype(str).str.strip().str.lower() == "true") &
        (app_df[pwd_col].astype(str).str.strip().str.lower() == "false") &
        (app_df["Service/générique?"].astype(str).str.upper().str.strip() == "FAUX") &
        (app_df["Compte Expiré?"].astype(str).str.upper().str.strip() == "NON") &
        (app_df["Comptes ayant une licence OneDrive"].astype(str).str.upper().str.strip() == "OUI") &
        (app_df["Comptes Sauvegardés"].astype(str).str.upper().str.strip() == "NON") &
        (app_df[">= 30 Jrs"].astype(str).str.upper().str.strip() == "NOK")
    )

    # Colonnes à afficher (enrichies)
    display_cols = [
        cn_col,
        upn_col,
        desc_col,
        dept_col,
        title_col,
        office_col,
        enabled_col,
        pwd_col,
        "Service/générique?",
        "Compte Expiré?",
        "Comptes ayant une licence OneDrive",
        "Comptes Sauvegardés",
        "Date d'obtention de la licence",
        "Date dernière synchronisation",
        ">= 30 Jrs"
    ]
    # Filtrer uniquement les colonnes existantes
    display_cols = [c for c in display_cols if c in app_df.columns]

    view = app_df.loc[mask, display_cols].copy().sort_values(by=[cn_col if cn_col in app_df.columns else desc_col], kind="stable", na_position="last")

    # ------ Stats ------
    # 1) Nombre d'utilisateurs actifs:
    cn_col = find_col(app_df, ["CN", "Name", "DisplayName"]) or "CN"
    if cn_col not in app_df.columns:
        app_df[cn_col] = ""

    actifs_mask = (
        (app_df[enabled_col].astype(str).str.strip().str.lower() == "true") &
        (app_df[cn_col].astype(str).str.strip() != "") &
        (app_df[cn_col].notna())
    )
    nb_utilisateurs_actifs = int(actifs_mask.sum())

    # 2) Nombre d'utilisateurs Assujettis:
    assujettis_mask = (
        (app_df[enabled_col].astype(str).str.strip().str.lower() == "true") &
        (app_df[pwd_col].astype(str).str.strip().str.lower() == "false") &
        (app_df["Service/générique?"].astype(str).str.upper().str.strip() == "FAUX") &
        (app_df["Compte Expiré?"].astype(str).str.upper().str.strip() == "NON")
    )
    nb_utilisateurs_assujettis = int(assujettis_mask.sum())

    # 3) Nombre Avec licence:
    avec_licence_mask = assujettis_mask & (
        app_df["Comptes ayant une licence OneDrive"].astype(str).str.upper().str.strip() == "OUI"
    )
    nb_avec_licence = int(avec_licence_mask.sum())

    # 4) Nombre d'utilisateurs NOK (utilisateurs qui n'ont pas synchronisé):
    nb_utilisateurs_nok = int(mask.sum())

    # 5) Taux : 1 - (nombre NOK / nombre avec licence)
    if nb_avec_licence > 0:
        taux = (1 - (nb_utilisateurs_nok / nb_avec_licence)) * 100
    else:
        taux = 0.0

    return [{
        "title": "Sauvegardes des données PCs",
        "excel_output": out_xlsx,
        "dataframe": dataframe_for_ui(view),
        "display_columns": [{"key": c, "label": c} for c in view.columns],
        "summary_stats": {
            "Les utilisateurs actifs": nb_utilisateurs_actifs,
            "Les utilisateurs Assujettis": nb_utilisateurs_assujettis,
            "Les utilisateurs avec licence": nb_avec_licence,
            "Les utilisateurs NOK": nb_utilisateurs_nok,
            "Taux": f"{round(taux, 2)}%"
        }
    }]

# ================================
# Entrée Hyper
# ================================
def run(input_file_paths: dict, output_dir_path: str):
    return traiter(
        onedrive_path=input_file_paths.get("onedrive_file"),
        users_path=input_file_paths.get("users_file"),
        ad_path=input_file_paths.get("ad_file"),
        output_dir=output_dir_path
    )